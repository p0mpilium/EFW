import pandas as pd
import datetime
from openpyxl import load_workbook

def read_excel_as_text(file_path, sheet_name):
    # Используем openpyxl для чтения листа целиком
    workbook = load_workbook(filename=file_path, data_only=True)
    sheet = workbook[sheet_name]

    full_text = []
    for row in sheet.iter_rows(values_only=True):
        row_text = ' '.join([str(cell) if cell is not None else '' for cell in row]).strip()
        if row_text:  # добавляем в список только непустые строки
            full_text.append(row_text)

    return full_text

def save_to_excel(data, save_path):
    # Создаем DataFrame и записываем данные
    df = pd.DataFrame({'Data': data})
    df.to_excel(save_path, index=False)
    print(f"Данные записаны в файл: {save_path}")

# Ввод пути файла и имени листа пользователем
file_path = input("Введите полный путь к файлу для чтения: ").strip()
sheet_name = input("Введите имя листа в книге Excel: ").strip()

# Читаем данные и удаляем пустые строки
text_data = read_excel_as_text(file_path, sheet_name)

# Задаем путь для сохранения нового файла
save_path = f"C:/Users/naver/Downloads/ОбработанныеДанные_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

# Сохраняем данные в новый файл
save_to_excel(text_data, save_path)
